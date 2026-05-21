"""Parse the per-state RTO-level Vahan files.

Filename pattern:  ``{State_Name}_Rto_{ColDim}_Unified.xlsx``
(e.g. ``Andhra_Pradesh_Rto_MonthWise_Unified.xlsx``)

The internal layout is the same cross-tab shape as the State-level files
(messy multi-row headers, Indian-format numbers, trailing TOTAL column), so we
reuse the row-extraction logic from :mod:`core.parser`.

Output schema (long form), one row per (rto, col_value):

    file_key, state, rto, col_dim, col_value, year, month, value

Where ``file_key`` is e.g. ``rto_monthwise``, ``rto_vehiclecategory`` —
the same key for every state, so all state files of the same column-dim
union into a single fact slice.
"""
from __future__ import annotations
import re
from pathlib import Path
from typing import Iterable
import openpyxl
import pandas as pd

from .config import COL_DIMS
from .parser import _norm, _to_number, _title_year, _find_header_and_data, _year_month_for

# Strict match: anything ending with "_Rto_{ColDim}_Unified.xlsx".
_RTO_FILENAME_RE = re.compile(r"^(.+?)_Rto_([A-Za-z]+)_Unified\.xlsx$", re.I)


def parse_rto_filename(path: Path) -> tuple[str, str] | None:
    """Return (state_name, col_dim) parsed from filename, or None."""
    m = _RTO_FILENAME_RE.match(path.name)
    if not m:
        return None
    state_raw, col = m.group(1), m.group(2)
    col_dim = next((d for d in COL_DIMS if d.lower() == col.lower()), None)
    if not col_dim:
        return None
    state = state_raw.replace("_", " ").strip()
    return state, col_dim


def parse_rto_file(path: Path) -> pd.DataFrame:
    """Parse one per-state RTO unified xlsx into long-form rows."""
    parsed = parse_rto_filename(path)
    if parsed is None:
        return pd.DataFrame()
    state, col_dim = parsed

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    title_year = _title_year(rows)
    header, data_start = _find_header_and_data(rows, col_dim)
    if data_start < 0:
        return pd.DataFrame()

    n_cols = max((len(r) for r in rows[data_start:]), default=0)
    header = header + [""] * max(0, n_cols - len(header))
    col_labels = [_norm(header[j] or "") for j in range(2, n_cols)]

    records = []
    file_key = f"rto_{col_dim}".lower()
    for r in rows[data_start:]:
        if not r or len(r) < 3:
            continue
        rv = _norm(r[1])
        if not rv or rv.upper() == "TOTAL":
            continue
        for j, lbl in enumerate(col_labels, start=2):
            if j >= len(r):
                break
            val = _to_number(r[j])
            if val is None:
                continue
            if not lbl or lbl.upper() == "TOTAL":
                continue
            year, month = _year_month_for(col_dim, lbl, title_year)
            records.append({
                "file_key": file_key,
                "state": state,
                "rto": rv,
                "col_dim": col_dim,
                "col_value": lbl,
                "year": year,
                "month": month,
                "value": val,
            })
    return pd.DataFrame.from_records(records)


def parse_rto_directory(directory: Path) -> Iterable[tuple[Path, pd.DataFrame]]:
    """Yield (path, dataframe) for every recognized RTO xlsx in directory."""
    for p in sorted(directory.glob("*_Rto_*_Unified.xlsx")):
        if parse_rto_filename(p) is None:
            continue
        try:
            df = parse_rto_file(p)
        except Exception as e:  # noqa: BLE001
            print(f"[rto_parser] FAILED {p.name}: {e}")
            continue
        yield p, df


# ----- dim_rto loader -------------------------------------------------------

_DIM_HEADER_ALIASES = {
    "rto_name": ["RTO Name", "Rto Name", "RTO_Name", "RTO"],
    "state_name": ["State Name", "State_Name", "State"],
    "region": ["Region"],
    "tier": [
        "The Urban/Rural Proxy (Strategic)",
        "Urban_Rural_Proxy",
        "Urban/Rural Proxy",
        "Tier",
    ],
}


def _resolve_col(header: list[str], aliases: list[str]) -> int | None:
    norm = [(_norm(h) or "").lower() for h in header]
    for a in aliases:
        a_low = a.lower()
        for i, h in enumerate(norm):
            if h == a_low:
                return i
    return None


def load_dim_rto(path: Path) -> pd.DataFrame:
    """Read the Consolidated_RTO_Stratification_Matrix.xlsx into dim_rto rows.

    Returns columns: rto_name, state_name, region, urban_rural_proxy.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # Use the first sheet whose header row contains "RTO Name"-ish + "Region".
    chosen = None
    for sn in wb.sheetnames:
        ws = wb[sn]
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first:
            continue
        h = [(_norm(c) or "").lower() for c in first]
        if any("rto" in x for x in h) and any(x == "region" for x in h):
            chosen = sn
            break
    if chosen is None:
        wb.close()
        raise ValueError(f"No usable sheet found in {path.name}")

    ws = wb[chosen]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return pd.DataFrame()

    header = list(rows[0])
    idx = {k: _resolve_col(header, v) for k, v in _DIM_HEADER_ALIASES.items()}

    records = []
    for r in rows[1:]:
        if not r or idx["rto_name"] is None:
            continue
        rto = _norm(r[idx["rto_name"]])
        if not rto:
            continue
        records.append({
            "rto_name": rto,
            "state_name": _norm(r[idx["state_name"]]) if idx["state_name"] is not None else None,
            "region": _norm(r[idx["region"]]) if idx["region"] is not None else None,
            "urban_rural_proxy": _norm(r[idx["tier"]]) if idx["tier"] is not None else None,
        })
    df = pd.DataFrame.from_records(records)
    df = df.replace({"": None})
    return df

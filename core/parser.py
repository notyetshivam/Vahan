"""Parse the 38 Vahan_*_*_Unified.xlsx files into a tidy long-form DataFrame.

Each file is a cross-tab of two dimensions (Dim1 on rows, Dim2 on columns) with
messy multi-row headers, nbsp padding, Indian-format number strings, and a
trailing TOTAL column. Output schema:

    file_key, row_dim, row_value, col_dim, col_value, year, month, value

Year semantics
--------------
* CalendarYear column dim   -> year = int(col_value)
* FinancialYear column dim  -> year = int(col_value.split('-')[0])
* MonthWise column dim      -> year = year-from-title, month = col_value
* Other column dims         -> year = year-from-title (current snapshot)

A row with col_value == 'TOTAL' is dropped to avoid double-counting; the per-row
total is recoverable by SUM(value) over col_value.
"""
from __future__ import annotations
import re
from pathlib import Path
from typing import Iterable
import pandas as pd
import openpyxl

from .config import ROW_DIMS, COL_DIMS

_FILENAME_RE = re.compile(r"Vahan_([A-Za-z]+)_([A-Za-z]+)_Unified\.xlsx$", re.I)
_YEAR_RE = re.compile(r"\((\d{4})(?:-\d{4})?\)")
_NBSP_RE = re.compile(r"[ \s]+")


def _norm(s) -> str:
    if s is None:
        return ""
    return _NBSP_RE.sub(" ", str(s)).strip()


def _to_number(v) -> float | None:
    """Parse Indian-format numeric strings like '1,19,134' or '-' into floats."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if s in {"", "-", "NA", "N/A", "null"}:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_filename(path: Path) -> tuple[str, str] | None:
    """Return (row_dim, col_dim) from a unified xlsx filename, or None."""
    m = _FILENAME_RE.search(path.name)
    if not m:
        return None
    a, b = m.group(1), m.group(2)
    # Validate against known dimensions (case-insensitive).
    row = next((d for d in ROW_DIMS if d.lower() == a.lower()), None)
    col = next((d for d in COL_DIMS if d.lower() == b.lower()), None)
    if not row or not col:
        return None
    return row, col


def _title_year(rows: list[list]) -> int | None:
    if not rows:
        return None
    title = _norm(rows[0][0]) if rows[0] else ""
    m = _YEAR_RE.search(title)
    return int(m.group(1)) if m else None


def _find_header_and_data(rows: list[list], col_dim: str) -> tuple[list[str], int]:
    """Locate the row that holds per-column labels and the first data row index.

    Standard layout: header in row index 3 (0-based), data from row 4.
    VehicleCategoryGroup column dim: header in row 4, data from row 5.
    Falls back to scanning for the first row whose 3rd cell can be cast to number.
    """
    # Find the first row where col C (index 2) is numeric-looking AND col B is a non-empty label.
    data_start = None
    for i, r in enumerate(rows):
        if i < 2:
            continue
        b = _norm(r[1]) if len(r) > 1 else ""
        c = r[2] if len(r) > 2 else None
        if b and _to_number(c) is not None:
            data_start = i
            break
    if data_start is None:
        return [], -1
    header_row_idx = data_start - 1
    header = [_norm(x) for x in rows[header_row_idx]]
    return header, data_start


def parse_file(path: Path) -> pd.DataFrame:
    """Parse one Vahan unified xlsx into long-form rows."""
    parsed = parse_filename(path)
    if parsed is None:
        return pd.DataFrame()
    row_dim, col_dim = parsed

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    title_year = _title_year(rows)
    header, data_start = _find_header_and_data(rows, col_dim)
    if data_start < 0:
        return pd.DataFrame()

    # Data columns: index 0 = S No, 1 = row_value, 2..N-2 = col values, N-1 = TOTAL.
    n_cols = max((len(r) for r in rows[data_start:]), default=0)
    # Pad header to width.
    header = header + [""] * max(0, n_cols - len(header))
    col_labels = []
    for j in range(2, n_cols):
        lbl = header[j] or ""
        col_labels.append(lbl)

    records = []
    for r in rows[data_start:]:
        if not r or len(r) < 3:
            continue
        rv = _norm(r[1])
        if not rv or rv.upper() == "TOTAL":
            continue
        for j, lbl in enumerate(col_labels, start=2):
            if j >= len(r):
                break
            val = _to_number(r[j])
            if val is None:
                continue
            cv_clean = _norm(lbl)
            if not cv_clean or cv_clean.upper() == "TOTAL":
                continue
            year, month = _year_month_for(col_dim, cv_clean, title_year)
            records.append({
                "file_key": f"{row_dim}_{col_dim}".lower(),
                "row_dim": row_dim,
                "row_value": rv,
                "col_dim": col_dim,
                "col_value": cv_clean,
                "year": year,
                "month": month,
                "value": val,
            })
    return pd.DataFrame.from_records(records)


def _year_month_for(col_dim: str, col_value: str, title_year: int | None):
    if col_dim == "CalendarYear":
        try:
            return int(col_value), None
        except ValueError:
            return title_year, None
    if col_dim == "FinancialYear":
        try:
            return int(col_value.split("-")[0]), None
        except (ValueError, IndexError):
            return title_year, None
    if col_dim == "MonthWise":
        return title_year, col_value.upper()
    return title_year, None


def parse_directory(directory: Path) -> Iterable[tuple[Path, pd.DataFrame]]:
    """Yield (path, dataframe) for every recognized unified xlsx in directory."""
    for p in sorted(directory.glob("Vahan_*_Unified.xlsx")):
        if parse_filename(p) is None:
            continue
        try:
            df = parse_file(p)
        except Exception as e:  # noqa: BLE001
            print(f"[parser] FAILED {p.name}: {e}")
            continue
        yield p, df
